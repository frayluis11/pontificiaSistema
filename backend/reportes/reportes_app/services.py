"""
Services para el microservicio de reportes

Contiene toda la lógica de negocio para generación de reportes,
procesamiento de métricas y exportación a diferentes formatos.
"""

import os
import json
import tempfile
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional, Union
from io import BytesIO

# Django imports
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db import transaction
from django.template.loader import render_to_string

# Third party imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

import matplotlib
matplotlib.use('Agg')  # Backend sin GUI
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

from .models import (
    Reporte, Metrica, ReporteMetrica, TipoReporte, EstadoReporte,
    FormatoReporte, PeriodoMetrica
)
from .repositories import ReportesRepositoryManager
from .exceptions import ReporteServiceException


class ReporteService:
    """
    Servicio principal para generación y gestión de reportes
    """
    
    def __init__(self):
        self.repositories = ReportesRepositoryManager()
        self.temp_dir = settings.REPORTES_CONFIG['TEMP_DIR']
        self.reports_dir = settings.REPORTES_CONFIG['REPORTS_DIR']
    
    def crear_reporte(self, nombre: str, tipo: str, parametros: dict, 
                     autor_id: str, autor_nombre: str, **kwargs) -> Reporte:
        """
        Crea un nuevo reporte
        """
        try:
            # Validar parámetros según el tipo
            self._validar_parametros_reporte(tipo, parametros)
            
            # Crear reporte
            reporte = self.repositories.reportes.crear(
                nombre=nombre,
                tipo=tipo,
                parametros=parametros,
                autor_id=autor_id,
                autor_nombre=autor_nombre,
                formato=kwargs.get('formato', FormatoReporte.PDF),
                descripcion=kwargs.get('descripcion', ''),
                autor_email=kwargs.get('autor_email', ''),
                notificar_por_email=kwargs.get('notificar_por_email', True),
                publico=kwargs.get('publico', False),
                fecha_expiracion=kwargs.get('fecha_expiracion'),
                tags=kwargs.get('tags', [])
            )
            
            return reporte
            
        except Exception as e:
            raise ReporteServiceException(f"Error creando reporte: {str(e)}")
    
    def generar_reporte_sync(self, reporte_id: str) -> Reporte:
        """
        Genera un reporte de forma síncrona
        """
        try:
            # Obtener y marcar reporte como procesando
            reporte = self.repositories.reportes.marcar_procesando(reporte_id)
            if not reporte:
                raise ReporteServiceException("Reporte no encontrado o no se puede procesar")
            
            # Generar contenido según el tipo
            datos = self._obtener_datos_reporte(reporte)
            
            # Generar archivo según el formato
            archivo_path = self._generar_archivo(reporte, datos)
            
            # Guardar archivo y completar reporte
            self._guardar_archivo_reporte(reporte, archivo_path)
            
            # Marcar como completado
            registros_procesados = len(datos.get('registros', []))
            self.repositories.reportes.completar_reporte(
                reporte_id, archivo_path, registros_procesados
            )
            
            return reporte
            
        except Exception as e:
            # Marcar reporte con error
            self.repositories.reportes.marcar_error(reporte_id, str(e))
            raise ReporteServiceException(f"Error generando reporte: {str(e)}")
    
    def generar_reporte_async(self, reporte_id: str):
        """
        Inicia generación asíncrona de reporte usando Celery
        """
        from .tasks import generar_reporte_async
        
        try:
            # Marcar como procesando
            reporte = self.repositories.reportes.marcar_procesando(reporte_id)
            if not reporte:
                raise ReporteServiceException("Reporte no encontrado")
            
            # Lanzar tarea asíncrona
            task = generar_reporte_async.delay(reporte_id)
            
            # Actualizar reporte con ID de tarea
            self.repositories.reportes.actualizar(
                reporte_id,
                parametros={**reporte.parametros, 'celery_task_id': task.id}
            )
            
            return {'task_id': task.id, 'reporte': reporte}
            
        except Exception as e:
            raise ReporteServiceException(f"Error iniciando generación asíncrona: {str(e)}")
    
    def _validar_parametros_reporte(self, tipo: str, parametros: dict):
        """Valida parámetros según el tipo de reporte"""
        
        validaciones_base = {
            'fecha_inicio': {'requerido': False, 'tipo': 'date'},
            'fecha_fin': {'requerido': False, 'tipo': 'date'},
            'formato_salida': {'requerido': False, 'valores': ['PDF', 'EXCEL', 'CSV']},
        }
        
        validaciones_especificas = {
            TipoReporte.ASISTENCIA: {
                'trabajador_ids': {'requerido': False, 'tipo': 'list'},
                'departamento': {'requerido': False, 'tipo': 'string'},
                'incluir_tardanzas': {'requerido': False, 'tipo': 'boolean'},
            },
            TipoReporte.PAGOS: {
                'planilla_ids': {'requerido': False, 'tipo': 'list'},
                'estado_pagos': {'requerido': False, 'valores': ['CALCULADO', 'APROBADO', 'PAGADO']},
                'incluir_descuentos': {'requerido': False, 'tipo': 'boolean'},
            },
            TipoReporte.METRICAS: {
                'metrica_nombres': {'requerido': True, 'tipo': 'list'},
                'periodo': {'requerido': False, 'valores': [p.value for p in PeriodoMetrica]},
                'incluir_graficos': {'requerido': False, 'tipo': 'boolean'},
            }
        }
        
        # Validar parámetros base
        self._aplicar_validaciones(parametros, validaciones_base)
        
        # Validar parámetros específicos
        if tipo in validaciones_especificas:
            self._aplicar_validaciones(parametros, validaciones_especificas[tipo])
    
    def _aplicar_validaciones(self, parametros: dict, validaciones: dict):
        """Aplica validaciones a los parámetros"""
        for campo, reglas in validaciones.items():
            valor = parametros.get(campo)
            
            # Validar requerido
            if reglas.get('requerido', False) and not valor:
                raise ReporteServiceException(f"El campo '{campo}' es requerido")
            
            if valor is not None:
                # Validar tipo
                tipo_esperado = reglas.get('tipo')
                if tipo_esperado == 'date' and not isinstance(valor, (date, str)):
                    raise ReporteServiceException(f"El campo '{campo}' debe ser una fecha")
                elif tipo_esperado == 'list' and not isinstance(valor, list):
                    raise ReporteServiceException(f"El campo '{campo}' debe ser una lista")
                elif tipo_esperado == 'boolean' and not isinstance(valor, bool):
                    raise ReporteServiceException(f"El campo '{campo}' debe ser booleano")
                elif tipo_esperado == 'string' and not isinstance(valor, str):
                    raise ReporteServiceException(f"El campo '{campo}' debe ser texto")
                
                # Validar valores permitidos
                valores_permitidos = reglas.get('valores')
                if valores_permitidos and valor not in valores_permitidos:
                    raise ReporteServiceException(
                        f"El campo '{campo}' debe ser uno de: {', '.join(valores_permitidos)}"
                    )
    
    def _obtener_datos_reporte(self, reporte: Reporte) -> dict:
        """Obtiene los datos necesarios para generar el reporte"""
        
        if reporte.tipo == TipoReporte.ASISTENCIA:
            return self._obtener_datos_asistencia(reporte.parametros)
        elif reporte.tipo == TipoReporte.PAGOS:
            return self._obtener_datos_pagos(reporte.parametros)
        elif reporte.tipo == TipoReporte.METRICAS:
            return self._obtener_datos_metricas(reporte.parametros)
        elif reporte.tipo == TipoReporte.ESTUDIANTES:
            return self._obtener_datos_estudiantes(reporte.parametros)
        elif reporte.tipo == TipoReporte.FINANCIERO:
            return self._obtener_datos_financiero(reporte.parametros)
        else:
            return self._obtener_datos_personalizado(reporte.parametros)
    
    def _obtener_datos_asistencia(self, parametros: dict) -> dict:
        """Obtiene datos de asistencia (placeholder - integración con microservicio de asistencia)"""
        # TODO: Integrar con API de asistencia
        return {
            'titulo': 'Reporte de Asistencia',
            'registros': [
                {
                    'trabajador': 'Juan Pérez',
                    'fecha': '2024-01-15',
                    'hora_entrada': '08:00',
                    'hora_salida': '17:00',
                    'estado': 'PRESENTE'
                },
                # ... más registros de ejemplo
            ],
            'resumen': {
                'total_trabajadores': 150,
                'presentes': 145,
                'ausentes': 5,
                'tardanzas': 10
            }
        }
    
    def _obtener_datos_pagos(self, parametros: dict) -> dict:
        """Obtiene datos de pagos (placeholder - integración con microservicio de pagos)"""
        # TODO: Integrar con API de pagos
        return {
            'titulo': 'Reporte de Pagos',
            'registros': [
                {
                    'trabajador': 'María González',
                    'monto_bruto': 3500.00,
                    'descuentos': 350.00,
                    'monto_neto': 3150.00,
                    'estado': 'PAGADO'
                },
                # ... más registros de ejemplo
            ],
            'resumen': {
                'total_pagos': 150,
                'monto_total_bruto': 525000.00,
                'monto_total_descuentos': 52500.00,
                'monto_total_neto': 472500.00
            }
        }
    
    def _obtener_datos_metricas(self, parametros: dict) -> dict:
        """Obtiene datos de métricas"""
        metrica_nombres = parametros.get('metrica_nombres', [])
        fecha_inicio = parametros.get('fecha_inicio')
        fecha_fin = parametros.get('fecha_fin')
        
        # Convertir fechas string a date si es necesario
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        if isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        datos_metricas = []
        
        for nombre in metrica_nombres:
            if fecha_inicio and fecha_fin:
                metricas = self.repositories.metricas.buscar_por_periodo(
                    'DIARIO', fecha_inicio, fecha_fin
                ).filter(nombre=nombre)
            else:
                metricas = self.repositories.metricas.obtener_metricas_recientes(nombre, 30)
            
            estadisticas = None
            if fecha_inicio and fecha_fin:
                estadisticas = self.repositories.metricas.calcular_estadisticas_metrica(
                    nombre, fecha_inicio, fecha_fin
                )
            
            datos_metricas.append({
                'nombre': nombre,
                'valores': [
                    {
                        'fecha': m.fecha_inicio,
                        'valor': float(m.valor),
                        'unidad': m.unidad
                    } for m in metricas
                ],
                'estadisticas': estadisticas
            })
        
        return {
            'titulo': 'Reporte de Métricas',
            'metricas': datos_metricas,
            'periodo': f"{fecha_inicio} - {fecha_fin}" if fecha_inicio and fecha_fin else "Últimos 30 registros"
        }
    
    def _obtener_datos_estudiantes(self, parametros: dict) -> dict:
        """Obtiene datos de estudiantes (placeholder)"""
        return {
            'titulo': 'Reporte de Estudiantes',
            'registros': [],
            'resumen': {}
        }
    
    def _obtener_datos_financiero(self, parametros: dict) -> dict:
        """Obtiene datos financieros (placeholder)"""
        return {
            'titulo': 'Reporte Financiero',
            'registros': [],
            'resumen': {}
        }
    
    def _obtener_datos_personalizado(self, parametros: dict) -> dict:
        """Obtiene datos para reporte personalizado"""
        return {
            'titulo': 'Reporte Personalizado',
            'registros': [],
            'resumen': {}
        }
    
    def _generar_archivo(self, reporte: Reporte, datos: dict) -> str:
        """Genera el archivo del reporte según el formato"""
        
        if reporte.formato == FormatoReporte.PDF:
            return self._generar_pdf(reporte, datos)
        elif reporte.formato == FormatoReporte.EXCEL:
            return self._generar_excel(reporte, datos)
        elif reporte.formato == FormatoReporte.CSV:
            return self._generar_csv(reporte, datos)
        elif reporte.formato == FormatoReporte.JSON:
            return self._generar_json(reporte, datos)
        else:
            raise ReporteServiceException(f"Formato no soportado: {reporte.formato}")
    
    def _generar_pdf(self, reporte: Reporte, datos: dict) -> str:
        """Genera reporte en formato PDF usando ReportLab"""
        
        # Crear archivo temporal
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{reporte.nombre}_{timestamp}.pdf"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Crear documento PDF
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centro
        )
        
        # Contenido del PDF
        story = []
        
        # Título
        title = Paragraph(datos.get('titulo', reporte.nombre), title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Información del reporte
        info_data = [
            ['Tipo:', reporte.tipo],
            ['Generado por:', reporte.autor_nombre],
            ['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Contenido específico según el tipo
        if reporte.tipo == TipoReporte.METRICAS:
            story.extend(self._generar_contenido_pdf_metricas(datos))
        else:
            # Contenido genérico para otros tipos
            story.extend(self._generar_contenido_pdf_generico(datos))
        
        # Construir PDF
        doc.build(story)
        
        return filepath
    
    def _generar_contenido_pdf_metricas(self, datos: dict) -> List:
        """Genera contenido específico para reportes de métricas en PDF"""
        story = []
        styles = getSampleStyleSheet()
        
        for metrica_data in datos.get('metricas', []):
            # Título de la métrica
            metrica_title = Paragraph(
                f"<b>{metrica_data['nombre']}</b>",
                styles['Heading2']
            )
            story.append(metrica_title)
            story.append(Spacer(1, 12))
            
            # Estadísticas si existen
            if metrica_data.get('estadisticas'):
                stats = metrica_data['estadisticas']
                stats_data = [
                    ['Promedio:', f"{stats.get('promedio', 0):.2f}"],
                    ['Máximo:', f"{stats.get('maximo', 0):.2f}"],
                    ['Mínimo:', f"{stats.get('minimo', 0):.2f}"],
                    ['Total registros:', str(stats.get('total', 0))],
                ]
                
                if stats.get('variacion_porcentual') is not None:
                    stats_data.append([
                        'Variación:', 
                        f"{stats['variacion_porcentual']:+.2f}%"
                    ])
                
                stats_table = Table(stats_data, colWidths=[2*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ]))
                
                story.append(stats_table)
            
            story.append(Spacer(1, 20))
        
        return story
    
    def _generar_contenido_pdf_generico(self, datos: dict) -> List:
        """Genera contenido genérico para otros tipos de reportes en PDF"""
        story = []
        styles = getSampleStyleSheet()
        
        # Resumen si existe
        if datos.get('resumen'):
            resumen_title = Paragraph("<b>Resumen</b>", styles['Heading2'])
            story.append(resumen_title)
            story.append(Spacer(1, 12))
            
            resumen_data = []
            for key, value in datos['resumen'].items():
                resumen_data.append([key.replace('_', ' ').title(), str(value)])
            
            if resumen_data:
                resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
                resumen_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                
                story.append(resumen_table)
                story.append(Spacer(1, 20))
        
        return story
    
    def _generar_excel(self, reporte: Reporte, datos: dict) -> str:
        """Genera reporte en formato Excel usando openpyxl"""
        
        # Crear archivo temporal
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{reporte.nombre}_{timestamp}.xlsx"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezado del reporte
        ws['A1'] = datos.get('titulo', reporte.nombre)
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Generado por: {reporte.autor_nombre}"
        ws['A3'] = f"Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Contenido específico según el tipo
        if reporte.tipo == TipoReporte.METRICAS:
            self._generar_contenido_excel_metricas(ws, datos, header_font, header_fill, border)
        else:
            self._generar_contenido_excel_generico(ws, datos, header_font, header_fill, border)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        wb.save(filepath)
        
        return filepath
    
    def _generar_contenido_excel_metricas(self, ws, datos, header_font, header_fill, border):
        """Genera contenido específico para métricas en Excel"""
        row = 5
        
        for metrica_data in datos.get('metricas', []):
            # Título de la métrica
            ws[f'A{row}'] = metrica_data['nombre']
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            # Encabezados de la tabla
            headers = ['Fecha', 'Valor', 'Unidad']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            row += 1
            
            # Datos de la métrica
            for valor_data in metrica_data.get('valores', []):
                ws.cell(row=row, column=1, value=valor_data.get('fecha', '')).border = border
                ws.cell(row=row, column=2, value=valor_data.get('valor', 0)).border = border
                ws.cell(row=row, column=3, value=valor_data.get('unidad', '')).border = border
                row += 1
            
            # Estadísticas si existen
            if metrica_data.get('estadisticas'):
                row += 1
                ws[f'A{row}'] = "Estadísticas:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                stats = metrica_data['estadisticas']
                for key, value in stats.items():
                    if key != 'valores' and value is not None:
                        ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                        ws.cell(row=row, column=2, value=f"{value:.2f}" if isinstance(value, (int, float)) else str(value))
                        row += 1
            
            row += 2
    
    def _generar_contenido_excel_generico(self, ws, datos, header_font, header_fill, border):
        """Genera contenido genérico para otros tipos de reportes en Excel"""
        row = 5
        
        # Resumen si existe
        if datos.get('resumen'):
            ws[f'A{row}'] = "Resumen"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            for key, value in datos['resumen'].items():
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            row += 2
        
        # Registros si existen
        registros = datos.get('registros', [])
        if registros:
            ws[f'A{row}'] = "Detalle"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            # Encabezados (basados en el primer registro)
            if registros:
                headers = list(registros[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                
                row += 1
                
                # Datos
                for registro in registros:
                    for col, key in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=registro.get(key, ''))
                        cell.border = border
                    row += 1
    
    def _generar_csv(self, reporte: Reporte, datos: dict) -> str:
        """Genera reporte en formato CSV"""
        import csv
        
        # Crear archivo temporal
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{reporte.nombre}_{timestamp}.csv"
        filepath = os.path.join(self.temp_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            # Para métricas
            if reporte.tipo == TipoReporte.METRICAS:
                writer = csv.writer(csvfile)
                writer.writerow(['Métrica', 'Fecha', 'Valor', 'Unidad'])
                
                for metrica_data in datos.get('metricas', []):
                    for valor_data in metrica_data.get('valores', []):
                        writer.writerow([
                            metrica_data['nombre'],
                            valor_data.get('fecha', ''),
                            valor_data.get('valor', 0),
                            valor_data.get('unidad', '')
                        ])
            
            # Para otros tipos
            else:
                registros = datos.get('registros', [])
                if registros:
                    headers = list(registros[0].keys())
                    writer = csv.DictWriter(csvfile, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(registros)
        
        return filepath
    
    def _generar_json(self, reporte: Reporte, datos: dict) -> str:
        """Genera reporte en formato JSON"""
        
        # Crear archivo temporal
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{reporte.nombre}_{timestamp}.json"
        filepath = os.path.join(self.temp_dir, filename)
        
        # Preparar datos para JSON
        json_data = {
            'reporte': {
                'id': str(reporte.id),
                'nombre': reporte.nombre,
                'tipo': reporte.tipo,
                'autor': reporte.autor_nombre,
                'fecha_generacion': timezone.now().isoformat(),
            },
            'datos': datos
        }
        
        # Convertir fechas y decimales para JSON
        def convert_for_json(obj):
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            elif isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False, default=convert_for_json)
        
        return filepath
    
    def _guardar_archivo_reporte(self, reporte: Reporte, archivo_path: str):
        """Guarda el archivo del reporte en el modelo"""
        try:
            with open(archivo_path, 'rb') as f:
                archivo_content = f.read()
            
            # Crear nombre final del archivo
            filename = os.path.basename(archivo_path)
            
            # Guardar en el modelo
            reporte.archivo.save(
                filename,
                ContentFile(archivo_content),
                save=True
            )
            
            # Actualizar información del archivo
            reporte.archivo_nombre = filename
            reporte.archivo_tamaño = len(archivo_content)
            reporte.save(update_fields=['archivo_nombre', 'archivo_tamaño'])
            
        except Exception as e:
            raise ReporteServiceException(f"Error guardando archivo: {str(e)}")
        finally:
            # Limpiar archivo temporal
            if os.path.exists(archivo_path):
                os.unlink(archivo_path)
    
    def cancelar_reporte(self, reporte_id: str) -> bool:
        """Cancela un reporte en procesamiento"""
        try:
            reporte = self.repositories.reportes.obtener_por_id(reporte_id)
            if reporte and reporte.estado in [EstadoReporte.PENDIENTE, EstadoReporte.PROCESANDO]:
                
                # Si tiene tarea de Celery, intentar cancelarla
                if 'celery_task_id' in reporte.parametros:
                    from celery import current_app
                    current_app.control.revoke(
                        reporte.parametros['celery_task_id'],
                        terminate=True
                    )
                
                reporte.cancelar()
                return True
            
            return False
            
        except Exception as e:
            raise ReporteServiceException(f"Error cancelando reporte: {str(e)}")
    
    def obtener_progreso_reporte(self, reporte_id: str) -> dict:
        """Obtiene el progreso de un reporte"""
        try:
            reporte = self.repositories.reportes.obtener_por_id(reporte_id)
            if not reporte:
                return None
            
            progreso = {
                'estado': reporte.estado,
                'fecha_inicio': reporte.fecha_inicio_procesamiento,
                'fecha_fin': reporte.fecha_fin_procesamiento,
                'registros_procesados': reporte.registros_procesados,
                'mensaje_error': reporte.mensaje_error,
                'tiempo_transcurrido': None
            }
            
            # Calcular tiempo transcurrido si está procesando
            if reporte.estado == EstadoReporte.PROCESANDO and reporte.fecha_inicio_procesamiento:
                tiempo_transcurrido = timezone.now() - reporte.fecha_inicio_procesamiento
                progreso['tiempo_transcurrido'] = tiempo_transcurrido.total_seconds()
            
            # Si hay tarea de Celery, obtener información adicional
            if 'celery_task_id' in reporte.parametros:
                from celery.result import AsyncResult
                task = AsyncResult(reporte.parametros['celery_task_id'])
                progreso['celery_estado'] = task.status
                if task.info:
                    progreso['celery_info'] = task.info
            
            return progreso
            
        except Exception as e:
            raise ReporteServiceException(f"Error obteniendo progreso: {str(e)}")


class MetricaService:
    """
    Servicio para gestión y cálculo de métricas
    """
    
    def __init__(self):
        self.repositories = ReportesRepositoryManager()
    
    def crear_metrica(self, nombre: str, valor: float, fecha_inicio: date,
                     fecha_fin: date, **kwargs) -> Metrica:
        """Crea una nueva métrica"""
        try:
            return self.repositories.metricas.crear(
                nombre=nombre,
                valor=Decimal(str(valor)),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                descripcion=kwargs.get('descripcion', ''),
                tipo=kwargs.get('tipo', 'CONTADOR'),
                periodo=kwargs.get('periodo', 'DIARIO'),
                unidad=kwargs.get('unidad', ''),
                categoria=kwargs.get('categoria', ''),
                etiquetas=kwargs.get('etiquetas', {}),
                fuente=kwargs.get('fuente', ''),
                autor_calculo_id=kwargs.get('autor_calculo_id', ''),
                valor_objetivo=kwargs.get('valor_objetivo'),
                umbral_minimo=kwargs.get('umbral_minimo'),
                umbral_maximo=kwargs.get('umbral_maximo')
            )
        except Exception as e:
            raise ReporteServiceException(f"Error creando métrica: {str(e)}")
    
    def calcular_metricas_periodo(self, fecha_inicio: date, fecha_fin: date) -> dict:
        """Calcula métricas para un período específico"""
        try:
            # Ejemplo de cálculos de métricas del sistema
            metricas_calculadas = {}
            
            # Métrica de reportes generados
            reportes_periodo = self.repositories.reportes.buscar_por_fecha_rango(
                fecha_inicio, fecha_fin
            )
            
            metricas_calculadas['reportes_generados'] = self.crear_metrica(
                nombre='Reportes Generados',
                valor=reportes_periodo.count(),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                categoria='Sistema',
                unidad='reportes',
                descripcion='Número total de reportes generados en el período'
            )
            
            # Métrica de reportes completados exitosamente
            reportes_completados = reportes_periodo.filter(
                estado=EstadoReporte.COMPLETADO
            )
            
            metricas_calculadas['reportes_completados'] = self.crear_metrica(
                nombre='Reportes Completados',
                valor=reportes_completados.count(),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                categoria='Sistema',
                unidad='reportes'
            )
            
            # Tasa de éxito
            if reportes_periodo.count() > 0:
                tasa_exito = (reportes_completados.count() / reportes_periodo.count()) * 100
                metricas_calculadas['tasa_exito_reportes'] = self.crear_metrica(
                    nombre='Tasa de Éxito de Reportes',
                    valor=tasa_exito,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    categoria='Calidad',
                    unidad='%',
                    tipo='PORCENTAJE'
                )
            
            return metricas_calculadas
            
        except Exception as e:
            raise ReporteServiceException(f"Error calculando métricas: {str(e)}")
    
    def exportar_metricas(self, metrica_nombres: List[str], formato: str = 'JSON',
                         fecha_inicio: date = None, fecha_fin: date = None) -> str:
        """Exporta métricas a diferentes formatos"""
        try:
            # Obtener métricas
            metricas_data = []
            
            for nombre in metrica_nombres:
                if fecha_inicio and fecha_fin:
                    metricas = self.repositories.metricas.buscar_por_periodo(
                        'DIARIO', fecha_inicio, fecha_fin
                    ).filter(nombre=nombre)
                else:
                    metricas = self.repositories.metricas.obtener_metricas_recientes(nombre)
                
                for metrica in metricas:
                    metricas_data.append({
                        'nombre': metrica.nombre,
                        'valor': float(metrica.valor),
                        'fecha_inicio': metrica.fecha_inicio,
                        'fecha_fin': metrica.fecha_fin,
                        'unidad': metrica.unidad,
                        'categoria': metrica.categoria,
                        'fecha_calculo': metrica.fecha_calculo
                    })
            
            # Crear archivo temporal
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            temp_dir = settings.REPORTES_CONFIG['TEMP_DIR']
            
            if formato.upper() == 'JSON':
                filename = f"metricas_export_{timestamp}.json"
                filepath = os.path.join(temp_dir, filename)
                
                def convert_for_json(obj):
                    if isinstance(obj, (datetime, date)):
                        return obj.isoformat()
                    elif isinstance(obj, Decimal):
                        return float(obj)
                    return obj
                
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(metricas_data, f, indent=2, ensure_ascii=False, 
                             default=convert_for_json)
            
            elif formato.upper() == 'CSV':
                import csv
                filename = f"metricas_export_{timestamp}.csv"
                filepath = os.path.join(temp_dir, filename)
                
                if metricas_data:
                    with open(filepath, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.DictWriter(f, fieldnames=metricas_data[0].keys())
                        writer.writeheader()
                        writer.writerows(metricas_data)
            
            elif formato.upper() == 'EXCEL':
                filename = f"metricas_export_{timestamp}.xlsx"
                filepath = os.path.join(temp_dir, filename)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Métricas"
                
                if metricas_data:
                    headers = list(metricas_data[0].keys())
                    
                    # Encabezados
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                    
                    # Datos
                    for row, metrica in enumerate(metricas_data, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=metrica[header])
                
                wb.save(filepath)
            
            else:
                raise ReporteServiceException(f"Formato no soportado: {formato}")
            
            return filepath
            
        except Exception as e:
            raise ReporteServiceException(f"Error exportando métricas: {str(e)}")


# Crear instancias de servicios para uso global
reporte_service = ReporteService()
metrica_service = MetricaService()