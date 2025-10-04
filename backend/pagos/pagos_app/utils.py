"""
Utilidades para el microservicio de pagos

Funciones auxiliares para validaciones, generación de PDFs,
comunicación con otros microservicios, etc.
"""

import hashlib
import os
import requests
from typing import Optional, Dict, Any
from decimal import Decimal
from datetime import datetime
from django.conf import settings
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import inch
import logging

logger = logging.getLogger(__name__)


def validar_trabajador_activo(trabajador_id: str) -> bool:
    """
    Valida si un trabajador está activo consultando el microservicio de personal
    
    Args:
        trabajador_id: ID del trabajador
    
    Returns:
        bool: True si está activo, False si no
    """
    try:
        # URL del microservicio de personal (configurar en settings)
        personal_service_url = getattr(
            settings, 
            'PERSONAL_SERVICE_URL', 
            'http://localhost:3002'
        )
        
        response = requests.get(
            f"{personal_service_url}/api/trabajadores/{trabajador_id}/",
            timeout=5
        )
        
        if response.status_code == 200:
            data = response.json()
            return data.get('activo', False)
        
        logger.warning(f"No se pudo validar trabajador {trabajador_id}: {response.status_code}")
        return False
        
    except requests.RequestException as e:
        logger.error(f"Error conectando con servicio de personal: {str(e)}")
        # En caso de error de conexión, asumir que está activo para no bloquear
        return True
    except Exception as e:
        logger.error(f"Error validando trabajador {trabajador_id}: {str(e)}")
        return False


def obtener_datos_trabajador(trabajador_id: str) -> Optional[Dict[str, Any]]:
    """
    Obtiene datos completos de un trabajador del microservicio de personal
    
    Args:
        trabajador_id: ID del trabajador
    
    Returns:
        Dict con datos del trabajador o None si no se encuentra
    """
    try:
        personal_service_url = getattr(
            settings, 
            'PERSONAL_SERVICE_URL', 
            'http://localhost:3002'
        )
        
        response = requests.get(
            f"{personal_service_url}/api/trabajadores/{trabajador_id}/",
            timeout=10
        )
        
        if response.status_code == 200:
            return response.json()
        
        return None
        
    except Exception as e:
        logger.error(f"Error obteniendo datos del trabajador {trabajador_id}: {str(e)}")
        return None


def generar_pdf_boleta(pago) -> ContentFile:
    """
    Genera el archivo PDF de la boleta de pago
    
    Args:
        pago: Instancia del modelo Pago
    
    Returns:
        ContentFile: Archivo PDF generado
    """
    try:
        from io import BytesIO
        
        # Crear buffer para el PDF
        buffer = BytesIO()
        
        # Configurar el documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Obtener estilos
        styles = getSampleStyleSheet()
        
        # Crear estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=30
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_LEFT,
            spaceAfter=12
        )
        
        # Crear contenido del PDF
        story = []
        
        # Título
        story.append(Paragraph("BOLETA DE PAGO", title_style))
        story.append(Spacer(1, 12))
        
        # Información de la empresa (configurar en settings)
        empresa_info = [
            ["SISTEMA PONTIFICIA"],
            ["RUC: 20123456789"],
            ["Dirección: Av. Principal 123, Lima, Perú"],
            ["Teléfono: (01) 123-4567"]
        ]
        
        empresa_table = Table(empresa_info, colWidths=[4*inch])
        empresa_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(empresa_table)
        story.append(Spacer(1, 20))
        
        # Información del trabajador y periodo
        info_data = [
            ["Trabajador:", pago.trabajador_nombre],
            ["Documento:", pago.trabajador_documento],
            ["Periodo:", f"{pago.planilla.año}-{pago.planilla.mes:02d}"],
            ["Días trabajados:", str(pago.dias_trabajados)],
            ["Fecha de pago:", datetime.now().strftime("%d/%m/%Y")]
        ]
        
        info_table = Table(info_data, colWidths=[1.5*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Detalle de pagos
        detalle_data = [
            ["CONCEPTO", "HABERES", "DESCUENTOS"]
        ]
        
        # Salario base
        detalle_data.append(["Salario Base", f"S/ {pago.salario_base:.2f}", ""])
        
        # Bonificaciones
        for pago_bonificacion in pago.pagobonificacion_set.all():
            monto = pago_bonificacion.monto_aplicado
            detalle_data.append([
                pago_bonificacion.bonificacion.nombre, 
                f"S/ {monto:.2f}", 
                ""
            ])
        
        # Descuentos
        for pago_descuento in pago.pagodescuento_set.all():
            monto = pago_descuento.monto_aplicado
            detalle_data.append([
                pago_descuento.descuento.nombre, 
                "", 
                f"S/ {monto:.2f}"
            ])
        
        # Totales
        detalle_data.extend([
            ["", "", ""],
            ["TOTAL HABERES", f"S/ {pago.monto_bruto:.2f}", ""],
            ["TOTAL DESCUENTOS", "", f"S/ {pago.total_descuentos:.2f}"],
            ["NETO A PAGAR", f"S/ {pago.monto_neto:.2f}", ""]
        ])
        
        detalle_table = Table(detalle_data, colWidths=[2.5*inch, 1.25*inch, 1.25*inch])
        detalle_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Contenido
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            
            # Líneas
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Totales
            ('BACKGROUND', (0, -3), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
            
            # Neto a pagar
            ('BACKGROUND', (0, -1), (-1, -1), colors.green),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
        ]))
        
        story.append(detalle_table)
        story.append(Spacer(1, 30))
        
        # Observaciones si las hay
        if pago.observaciones:
            story.append(Paragraph("Observaciones:", header_style))
            story.append(Paragraph(pago.observaciones, styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Pie de página
        pie_data = [
            ["Este documento es generado automáticamente por el Sistema Pontificia"],
            [f"Boleta N°: {pago.boleta.numero_boleta if hasattr(pago, 'boleta') else 'N/A'}"],
            [f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"]
        ]
        
        pie_table = Table(pie_data, colWidths=[5*inch])
        pie_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
        ]))
        
        story.append(pie_table)
        
        # Construir PDF
        doc.build(story)
        
        # Obtener contenido del buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Crear nombre del archivo
        filename = f"boleta_{pago.trabajador_documento}_{pago.planilla.año}{pago.planilla.mes:02d}.pdf"
        
        # Retornar como ContentFile
        return ContentFile(pdf_content, name=filename)
        
    except Exception as e:
        logger.error(f"Error generando PDF de boleta para pago {pago.id}: {str(e)}")
        raise


def calcular_hash_archivo(archivo) -> str:
    """
    Calcula el hash SHA-256 de un archivo
    
    Args:
        archivo: Archivo a procesar
    
    Returns:
        str: Hash SHA-256 del archivo
    """
    hash_sha256 = hashlib.sha256()
    
    for chunk in archivo.chunks():
        hash_sha256.update(chunk)
    
    return hash_sha256.hexdigest()


def formatear_moneda(monto: Decimal) -> str:
    """
    Formatea un monto como moneda peruana
    
    Args:
        monto: Monto a formatear
    
    Returns:
        str: Monto formateado como "S/ XX.XX"
    """
    return f"S/ {monto:.2f}"


def numero_a_letras(numero: Decimal) -> str:
    """
    Convierte un número a su representación en letras (español)
    
    Args:
        numero: Número a convertir
    
    Returns:
        str: Número en letras
    """
    # Esta es una implementación básica
    # En un sistema real se usaría una librería como num2words
    
    unidades = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
    
    if numero == 0:
        return "CERO"
    
    # Para simplicidad, solo manejo números hasta 999
    if numero < 10:
        return unidades[int(numero)]
    
    return f"NÚMERO: {numero:.2f}"


def validar_periodo_valido(año: int, mes: int) -> bool:
    """
    Valida que un periodo sea válido
    
    Args:
        año: Año del periodo
        mes: Mes del periodo
    
    Returns:
        bool: True si es válido
    """
    if año < 2020 or año > 2030:
        return False
    
    if mes < 1 or mes > 12:
        return False
    
    return True


def obtener_configuracion_pago(clave: str, default=None):
    """
    Obtiene configuración específica de pagos
    
    Args:
        clave: Clave de configuración
        default: Valor por defecto
    
    Returns:
        Valor de configuración
    """
    try:
        config = getattr(settings, 'PAGOS_CONFIG', {})
        return config.get(clave, default)
    except:
        return default


def enviar_email_boleta(boleta, email_trabajador: str) -> bool:
    """
    Envía la boleta por email al trabajador
    
    Args:
        boleta: Instancia de Boleta
        email_trabajador: Email del trabajador
    
    Returns:
        bool: True si se envió exitosamente
    """
    try:
        from django.core.mail import EmailMessage
        
        subject = f"Boleta de Pago - {boleta.pago.planilla.año}-{boleta.pago.planilla.mes:02d}"
        message = f"""
        Estimado/a {boleta.pago.trabajador_nombre},
        
        Adjunto encontrará su boleta de pago correspondiente al periodo {boleta.pago.planilla.año}-{boleta.pago.planilla.mes:02d}.
        
        Detalles del pago:
        - Monto bruto: {formatear_moneda(boleta.pago.monto_bruto)}
        - Total descuentos: {formatear_moneda(boleta.pago.total_descuentos)}
        - Neto a pagar: {formatear_moneda(boleta.pago.monto_neto)}
        
        Saludos cordiales,
        Sistema Pontificia
        """
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_trabajador]
        )
        
        if boleta.archivo_pdf:
            email.attach_file(boleta.archivo_pdf.path)
        
        email.send()
        
        # Marcar como enviada
        boleta.enviada_por_email = True
        boleta.fecha_envio_email = datetime.now()
        boleta.save()
        
        logger.info(f"Boleta {boleta.numero_boleta} enviada por email a {email_trabajador}")
        return True
        
    except Exception as e:
        logger.error(f"Error enviando boleta por email: {str(e)}")
        return False


def generar_reporte_planilla_excel(planilla):
    """
    Genera reporte de planilla en formato Excel
    
    Args:
        planilla: Instancia de Planilla
    
    Returns:
        BytesIO: Archivo Excel generado
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Planilla {planilla.año}-{planilla.mes:02d}"
        
        # Encabezados
        headers = [
            'Trabajador', 'Documento', 'Salario Base', 'Bonificaciones',
            'Descuentos', 'Monto Bruto', 'Monto Neto', 'Estado'
        ]
        
        # Agregar encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Agregar datos
        for row, pago in enumerate(planilla.pagos.all(), 2):
            ws.cell(row=row, column=1, value=pago.trabajador_nombre)
            ws.cell(row=row, column=2, value=pago.trabajador_documento)
            ws.cell(row=row, column=3, value=float(pago.salario_base))
            ws.cell(row=row, column=4, value=float(pago.total_bonificaciones))
            ws.cell(row=row, column=5, value=float(pago.total_descuentos))
            ws.cell(row=row, column=6, value=float(pago.monto_bruto))
            ws.cell(row=row, column=7, value=float(pago.monto_neto))
            ws.cell(row=row, column=8, value=pago.estado)
        
        # Agregar totales
        total_row = len(planilla.pagos.all()) + 3
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(planilla.total_bruto)).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(planilla.total_neto)).font = Font(bold=True)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error generando reporte Excel de planilla {planilla.id}: {str(e)}")
        raise