"""
Service layer para el microservicio de auditoría
"""

from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
import logging
import json
import csv
import io
import tempfile
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

from .models import ActividadSistema
from .repositories import AuditoriaRepository

logger = logging.getLogger('auditoria_app')


class AuditoriaService:
    """
    Service para gestión de auditoría del sistema
    """
    
    def __init__(self):
        self.repository = AuditoriaRepository()
        self.cache_timeout = getattr(settings, 'AUDITORIA_CACHE_TIMEOUT', 300)
    
    # ========== REGISTROS DE ACTIVIDADES ==========
    
    def registrar_actividad(self, 
                           usuario_id: Optional[int] = None,
                           usuario_email: str = None,
                           accion: str = 'OTHER',
                           recurso: str = 'OTHER',
                           recurso_id: str = None,
                           detalle: Dict = None,
                           request_data: Dict = None,
                           **kwargs) -> ActividadSistema:
        """
        Registrar nueva actividad en el sistema
        
        Args:
            usuario_id: ID del usuario
            usuario_email: Email del usuario
            accion: Tipo de acción
            recurso: Recurso afectado
            recurso_id: ID del recurso
            detalle: Detalles adicionales
            request_data: Datos de la request HTTP
            **kwargs: Otros parámetros
        
        Returns:
            ActividadSistema: Actividad registrada
        """
        try:
            # Preparar datos base
            data = {
                'usuario_id': usuario_id,
                'usuario_email': usuario_email,
                'accion': accion,
                'recurso': recurso,
                'recurso_id': recurso_id,
                'detalle': detalle or {},
                **kwargs
            }
            
            # Agregar datos de la request si están disponibles
            if request_data:
                data.update({
                    'ip_address': request_data.get('ip_address'),
                    'user_agent': request_data.get('user_agent'),
                    'session_key': request_data.get('session_key'),
                    'metodo_http': request_data.get('method'),
                    'url_solicitada': request_data.get('url'),
                    'codigo_estado': request_data.get('status_code'),
                })
            
            # Determinar nivel de criticidad automáticamente
            if 'nivel_criticidad' not in data:
                data['nivel_criticidad'] = self._determinar_criticidad(accion, recurso, kwargs.get('exito', True))
            
            # Agregar tags automáticos
            tags = kwargs.get('tags', [])
            tags.extend(self._generar_tags_automaticos(accion, recurso))
            data['tags'] = list(set(tags))  # Eliminar duplicados
            
            actividad = self.repository.crear_actividad(**data)
            
            # Actualizar caché de estadísticas
            self._invalidar_cache_estadisticas()
            
            # Log adicional para actividades críticas
            if actividad.is_critical_activity():
                logger.warning(f"Actividad crítica registrada: {actividad}")
            
            return actividad
            
        except Exception as e:
            logger.error(f"Error al registrar actividad: {str(e)}")
            raise
    
    def registrar_login(self, usuario_id: int, usuario_email: str = None,
                       ip_address: str = None, user_agent: str = None,
                       exito: bool = True, detalle: Dict = None) -> ActividadSistema:
        """
        Registrar intento de login
        
        Args:
            usuario_id: ID del usuario
            usuario_email: Email del usuario
            ip_address: Dirección IP
            user_agent: User agent
            exito: Si el login fue exitoso
            detalle: Detalles adicionales
        
        Returns:
            ActividadSistema: Actividad registrada
        """
        accion = 'LOGIN' if exito else 'LOGIN_FAILED'
        nivel_criticidad = 'LOW' if exito else 'HIGH'
        
        detalle_login = detalle or {}
        detalle_login.update({
            'login_attempt': True,
            'success': exito,
            'timestamp': timezone.now().isoformat()
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            usuario_email=usuario_email,
            accion=accion,
            recurso='AUTH',
            detalle=detalle_login,
            ip_address=ip_address,
            user_agent=user_agent,
            nivel_criticidad=nivel_criticidad,
            exito=exito
        )
    
    def registrar_operacion_crud(self, usuario_id: int, accion: str,
                                recurso: str, recurso_id: str = None,
                                detalle: Dict = None, exito: bool = True,
                                **kwargs) -> ActividadSistema:
        """
        Registrar operación CRUD
        
        Args:
            usuario_id: ID del usuario
            accion: Tipo de operación (CREATE, READ, UPDATE, DELETE)
            recurso: Recurso afectado
            recurso_id: ID del recurso
            detalle: Detalles de la operación
            exito: Si la operación fue exitosa
            **kwargs: Otros parámetros
        
        Returns:
            ActividadSistema: Actividad registrada
        """
        # Nivel de criticidad según la operación
        criticidad_map = {
            'CREATE': 'LOW',
            'READ': 'LOW',
            'UPDATE': 'MEDIUM',
            'DELETE': 'HIGH'
        }
        
        nivel_criticidad = criticidad_map.get(accion.upper(), 'LOW')
        
        detalle_crud = detalle or {}
        detalle_crud.update({
            'operation_type': 'CRUD',
            'resource_affected': recurso,
            'resource_id': recurso_id
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion=accion.upper(),
            recurso=recurso,
            recurso_id=recurso_id,
            detalle=detalle_crud,
            nivel_criticidad=nivel_criticidad,
            exito=exito,
            **kwargs
        )
    
    def registrar_error(self, usuario_id: Optional[int], recurso: str,
                       mensaje_error: str, detalle: Dict = None,
                       **kwargs) -> ActividadSistema:
        """
        Registrar error del sistema
        
        Args:
            usuario_id: ID del usuario (puede ser None)
            recurso: Recurso donde ocurrió el error
            mensaje_error: Mensaje de error
            detalle: Detalles adicionales
            **kwargs: Otros parámetros
        
        Returns:
            ActividadSistema: Actividad registrada
        """
        detalle_error = detalle or {}
        detalle_error.update({
            'error_type': 'SYSTEM_ERROR',
            'error_message': mensaje_error,
            'timestamp': timezone.now().isoformat()
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion='ERROR',
            recurso=recurso,
            detalle=detalle_error,
            mensaje_error=mensaje_error,
            nivel_criticidad='HIGH',
            exito=False,
            **kwargs
        )
    
    # ========== CONSULTAS Y FILTROS ==========
    
    def obtener_actividades_usuario(self, usuario_id: int,
                                   limite: int = 50) -> List[ActividadSistema]:
        """
        Obtener actividades de un usuario
        
        Args:
            usuario_id: ID del usuario
            limite: Límite de resultados
        
        Returns:
            Lista de actividades del usuario
        """
        cache_key = f"actividades_usuario_{usuario_id}_{limite}"
        actividades = cache.get(cache_key)
        
        if actividades is None:
            actividades = list(self.repository.obtener_por_usuario(usuario_id, limite))
            cache.set(cache_key, actividades, self.cache_timeout)
        
        return actividades
    
    def obtener_actividades_filtradas(self, filtros: Dict[str, Any]) -> List[ActividadSistema]:
        """
        Obtener actividades con filtros avanzados
        
        Args:
            filtros: Diccionario con filtros
        
        Returns:
            Lista de actividades filtradas
        """
        return list(self.repository.obtener_filtros_avanzados(**filtros))
    
    def buscar_actividades(self, termino: str, campos: List[str] = None) -> List[ActividadSistema]:
        """
        Buscar actividades por término
        
        Args:
            termino: Término de búsqueda
            campos: Campos donde buscar
        
        Returns:
            Lista de actividades encontradas
        """
        return list(self.repository.buscar_actividades(termino, campos))
    
    def obtener_actividades_criticas(self, horas: int = 24) -> List[ActividadSistema]:
        """
        Obtener actividades críticas recientes
        
        Args:
            horas: Horas hacia atrás para buscar
        
        Returns:
            Lista de actividades críticas
        """
        fecha_inicio = timezone.now() - timedelta(hours=horas)
        return list(self.repository.obtener_actividades_criticas(fecha_inicio))
    
    # ========== ESTADÍSTICAS Y ANALYTICS ==========
    
    def obtener_estadisticas_sistema(self, fecha_inicio: datetime = None,
                                   fecha_fin: datetime = None) -> Dict[str, Any]:
        """
        Obtener estadísticas del sistema
        
        Args:
            fecha_inicio: Fecha de inicio
            fecha_fin: Fecha de fin
        
        Returns:
            Dict con estadísticas
        """
        cache_key = f"stats_sistema_{fecha_inicio}_{fecha_fin}"
        stats = cache.get(cache_key)
        
        if stats is None:
            stats = self.repository.obtener_estadisticas_sistema(fecha_inicio, fecha_fin)
            
            # Agregar métricas adicionales
            stats.update(self._calcular_metricas_adicionales(fecha_inicio, fecha_fin))
            
            cache.set(cache_key, stats, self.cache_timeout)
        
        return stats
    
    def obtener_estadisticas_usuario(self, usuario_id: int) -> Dict[str, Any]:
        """
        Obtener estadísticas de un usuario
        
        Args:
            usuario_id: ID del usuario
        
        Returns:
            Dict con estadísticas del usuario
        """
        cache_key = f"stats_usuario_{usuario_id}"
        stats = cache.get(cache_key)
        
        if stats is None:
            stats = self.repository.obtener_estadisticas_usuario(usuario_id)
            cache.set(cache_key, stats, self.cache_timeout)
        
        return stats
    
    def detectar_actividades_sospechosas(self, horas: int = 24) -> List[Dict[str, Any]]:
        """
        Detectar actividades sospechosas
        
        Args:
            horas: Horas hacia atrás para analizar
        
        Returns:
            Lista de actividades sospechosas
        """
        return self.repository.obtener_actividades_sospechosas(horas)
    
    def obtener_metricas_rendimiento(self) -> Dict[str, Any]:
        """
        Obtener métricas de rendimiento
        
        Returns:
            Dict con métricas de rendimiento
        """
        cache_key = "metricas_rendimiento"
        metricas = cache.get(cache_key)
        
        if metricas is None:
            metricas = self.repository.obtener_metricas_rendimiento()
            cache.set(cache_key, metricas, self.cache_timeout)
        
        return metricas
    
    # ========== EXPORTACIÓN ==========
    
    def exportar_actividades_csv(self, filtros: Dict[str, Any] = None) -> str:
        """
        Exportar actividades a CSV
        
        Args:
            filtros: Filtros a aplicar
        
        Returns:
            Ruta del archivo CSV generado
        """
        try:
            actividades = self.repository.obtener_filtros_avanzados(**(filtros or {}))
            
            # Crear archivo temporal
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Headers
                headers = [
                    'ID', 'Usuario ID', 'Usuario Email', 'Acción', 'Recurso',
                    'Recurso ID', 'Fecha', 'IP Address', 'Éxito', 'Criticidad',
                    'Microservicio', 'Detalles'
                ]
                writer.writerow(headers)
                
                # Datos
                for actividad in actividades:
                    row = [
                        str(actividad.id),
                        actividad.usuario_id or '',
                        actividad.usuario_email or '',
                        actividad.get_accion_display(),
                        actividad.get_recurso_display(),
                        actividad.recurso_id or '',
                        actividad.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                        actividad.ip_address or '',
                        'Sí' if actividad.exito else 'No',
                        actividad.get_nivel_criticidad_display(),
                        actividad.microservicio,
                        actividad.get_detalle_formatted()
                    ]
                    writer.writerow(row)
                
                logger.info(f"Archivo CSV generado: {f.name}")
                return f.name
                
        except Exception as e:
            logger.error(f"Error al exportar CSV: {str(e)}")
            raise
    
    def exportar_actividades_excel(self, filtros: Dict[str, Any] = None) -> str:
        """
        Exportar actividades a Excel
        
        Args:
            filtros: Filtros a aplicar
        
        Returns:
            Ruta del archivo Excel generado
        """
        try:
            actividades = self.repository.obtener_filtros_avanzados(**(filtros or {}))
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Actividades de Auditoría"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Headers
            headers = [
                'ID', 'Usuario ID', 'Usuario Email', 'Acción', 'Recurso',
                'Recurso ID', 'Fecha', 'IP Address', 'Éxito', 'Criticidad',
                'Microservicio', 'Duración (ms)', 'Detalles'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos
            for row, actividad in enumerate(actividades, 2):
                data = [
                    str(actividad.id),
                    actividad.usuario_id or '',
                    actividad.usuario_email or '',
                    actividad.get_accion_display(),
                    actividad.get_recurso_display(),
                    actividad.recurso_id or '',
                    actividad.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                    actividad.ip_address or '',
                    'Sí' if actividad.exito else 'No',
                    actividad.get_nivel_criticidad_display(),
                    actividad.microservicio,
                    actividad.duracion_ms or '',
                    actividad.get_detalle_formatted()
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                wb.save(f.name)
                logger.info(f"Archivo Excel generado: {f.name}")
                return f.name
                
        except Exception as e:
            logger.error(f"Error al exportar Excel: {str(e)}")
            raise
    
    def exportar_actividades_pdf(self, filtros: Dict[str, Any] = None) -> str:
        """
        Exportar actividades a PDF
        
        Args:
            filtros: Filtros a aplicar
        
        Returns:
            Ruta del archivo PDF generado
        """
        try:
            actividades = list(self.repository.obtener_filtros_avanzados(**(filtros or {}))[:100])  # Limitar para PDF
            
            # Crear archivo temporal
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
                doc = SimpleDocTemplate(f.name, pagesize=A4)
                story = []
                
                # Estilos
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    spaceAfter=30,
                    alignment=1  # Center
                )
                
                # Título
                title = Paragraph("Reporte de Auditoría - Actividades del Sistema", title_style)
                story.append(title)
                story.append(Spacer(1, 20))
                
                # Información del reporte
                info_text = f"""
                Fecha de generación: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
                Total de actividades: {len(actividades)}<br/>
                Período: {filtros.get('fecha_inicio', 'N/A')} - {filtros.get('fecha_fin', 'N/A')}
                """
                info = Paragraph(info_text, styles['Normal'])
                story.append(info)
                story.append(Spacer(1, 20))
                
                # Tabla de actividades
                if actividades:
                    data = [['Fecha', 'Usuario', 'Acción', 'Recurso', 'Éxito', 'Criticidad']]
                    
                    for actividad in actividades:
                        row = [
                            actividad.fecha.strftime('%Y-%m-%d %H:%M'),
                            actividad.usuario_email or f"ID: {actividad.usuario_id}" or 'N/A',
                            actividad.get_accion_display(),
                            actividad.get_recurso_display(),
                            'Sí' if actividad.exito else 'No',
                            actividad.get_nivel_criticidad_display()
                        ]
                        data.append(row)
                    
                    table = Table(data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                else:
                    no_data = Paragraph("No se encontraron actividades con los filtros especificados.", styles['Normal'])
                    story.append(no_data)
                
                # Generar PDF
                doc.build(story)
                logger.info(f"Archivo PDF generado: {f.name}")
                return f.name
                
        except Exception as e:
            logger.error(f"Error al exportar PDF: {str(e)}")
            raise
    
    # ========== UTILIDADES PRIVADAS ==========
    
    def _determinar_criticidad(self, accion: str, recurso: str, exito: bool) -> str:
        """Determinar nivel de criticidad automáticamente"""
        if not exito:
            return 'HIGH'
        
        critical_actions = ['DELETE', 'LOGIN_FAILED', 'ERROR', 'EXCEPTION', 'ACCESS_DENIED']
        high_actions = ['UPDATE', 'STATUS_CHANGE', 'APPROVE', 'REJECT']
        
        if accion in critical_actions:
            return 'CRITICAL' if accion in ['DELETE', 'ERROR'] else 'HIGH'
        elif accion in high_actions:
            return 'MEDIUM'
        else:
            return 'LOW'
    
    def _generar_tags_automaticos(self, accion: str, recurso: str) -> List[str]:
        """Generar tags automáticos basados en acción y recurso"""
        tags = []
        
        # Tags por acción
        if accion in ['LOGIN', 'LOGOUT', 'LOGIN_FAILED']:
            tags.append('authentication')
        elif accion in ['CREATE', 'READ', 'UPDATE', 'DELETE']:
            tags.append('crud')
        elif accion in ['ERROR', 'EXCEPTION']:
            tags.append('error')
        
        # Tags por recurso
        if recurso in ['DOCUMENTS', 'DOCUMENT_FILE']:
            tags.append('documents')
        elif recurso in ['PAYMENTS', 'PAYMENT_TRANSACTION']:
            tags.append('payments')
        elif recurso in ['AUTH', 'USER_PROFILE']:
            tags.append('users')
        
        return tags
    
    def _calcular_metricas_adicionales(self, fecha_inicio: datetime = None,
                                     fecha_fin: datetime = None) -> Dict[str, Any]:
        """Calcular métricas adicionales del sistema"""
        metricas = {}
        
        # Tasa de éxito
        filtros = {}
        if fecha_inicio:
            filtros['fecha__gte'] = fecha_inicio
        if fecha_fin:
            filtros['fecha__lte'] = fecha_fin
        
        total = self.repository.count(**filtros)
        exitosas = self.repository.count(exito=True, **filtros)
        
        metricas['tasa_exito'] = (exitosas / total * 100) if total > 0 else 0
        
        # Actividades por hora del día
        metricas['actividades_por_hora'] = self._obtener_actividades_por_hora(filtros)
        
        return metricas
    
    def _obtener_actividades_por_hora(self, filtros: Dict) -> Dict[int, int]:
        """Obtener distribución de actividades por hora"""
        from django.db.models import Count, Extract
        
        actividades = self.repository.filter(**filtros).annotate(
            hora=Extract('fecha', 'hour')
        ).values('hora').annotate(
            count=Count('hora')
        ).order_by('hora')
        
        return {item['hora']: item['count'] for item in actividades}
    
    def _invalidar_cache_estadisticas(self):
        """Invalidar caché de estadísticas"""
        cache_keys = [
            'stats_sistema_*',
            'metricas_rendimiento',
        ]
        for pattern in cache_keys:
            cache.delete_pattern(pattern)
    
    # ========== MANTENIMIENTO ==========
    
    def limpiar_actividades_antiguas(self, dias_retencion: int = None) -> int:
        """
        Limpiar actividades antiguas
        
        Args:
            dias_retencion: Días de retención (por defecto usa configuración)
        
        Returns:
            Número de actividades eliminadas
        """
        if dias_retencion is None:
            dias_retencion = getattr(settings, 'AUDITORIA_CONFIG', {}).get('RETENTION_DAYS', 365)
        
        eliminadas = self.repository.limpiar_actividades_antiguas(dias_retencion)
        
        # Invalidar caché después de la limpieza
        self._invalidar_cache_estadisticas()
        
        return eliminadas
    
    # ========== MÉTODOS PARA OBSERVER ==========
    
    def registrar_cambio_documento(self, usuario_id: int = None, documento_id: str = None,
                                  accion: str = 'UPDATE', detalle: Dict = None,
                                  microservicio: str = 'DOCUMENTS') -> ActividadSistema:
        """
        Registrar cambio en documento
        """
        detalle_doc = detalle or {}
        detalle_doc.update({
            'documento_id': documento_id,
            'microservicio': microservicio,
            'cambio_automatico': True
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion=accion,
            recurso='DOCUMENT_FILE',
            recurso_id=documento_id,
            detalle=detalle_doc,
            nivel_criticidad='MEDIUM'
        )
    
    def registrar_cambio_pago(self, usuario_id: int = None, pago_id: str = None,
                             accion: str = 'UPDATE', detalle: Dict = None,
                             microservicio: str = 'PAYMENTS') -> ActividadSistema:
        """
        Registrar cambio en pago
        """
        detalle_pago = detalle or {}
        detalle_pago.update({
            'pago_id': pago_id,
            'microservicio': microservicio,
            'cambio_automatico': True
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion=accion,
            recurso='PAYMENT_TRANSACTION',
            recurso_id=pago_id,
            detalle=detalle_pago,
            nivel_criticidad='HIGH'
        )
    
    def registrar_cambio_usuario(self, usuario_id: int = None, accion: str = 'UPDATE',
                                detalle: Dict = None, microservicio: str = 'USERS') -> ActividadSistema:
        """
        Registrar cambio en usuario
        """
        detalle_usuario = detalle or {}
        detalle_usuario.update({
            'microservicio': microservicio,
            'cambio_automatico': True
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion=accion,
            recurso='USER_PROFILE',
            recurso_id=str(usuario_id),
            detalle=detalle_usuario,
            nivel_criticidad='MEDIUM'
        )
    
    def registrar_cambio_asistencia(self, usuario_id: int = None, asistencia_id: str = None,
                                   accion: str = 'UPDATE', detalle: Dict = None,
                                   microservicio: str = 'ATTENDANCE') -> ActividadSistema:
        """
        Registrar cambio en asistencia
        """
        detalle_asistencia = detalle or {}
        detalle_asistencia.update({
            'asistencia_id': asistencia_id,
            'microservicio': microservicio,
            'cambio_automatico': True
        })
        
        return self.registrar_actividad(
            usuario_id=usuario_id,
            accion=accion,
            recurso='ATTENDANCE_RECORD',
            recurso_id=asistencia_id,
            detalle=detalle_asistencia,
            nivel_criticidad='LOW'
        )
    
    # ========== MÉTODOS DE EXPORTACIÓN ==========
    
    def exportar_csv(self, filtros: Dict = None) -> str:
        """
        Exportar logs a CSV
        """
        try:
            actividades = self.repository.obtener_actividades_filtradas(filtros or {})
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Escribir headers
            writer.writerow([
                'ID', 'Fecha', 'Usuario ID', 'Usuario Email', 'Acción',
                'Recurso', 'Recurso ID', 'IP Address', 'Código Estado',
                'Duración (ms)', 'Nivel Criticidad', 'Detalles'
            ])
            
            # Escribir datos
            for actividad in actividades:
                writer.writerow([
                    str(actividad.id),
                    actividad.fecha.strftime('%Y-%m-%d %H:%M:%S'),
                    actividad.usuario_id or '',
                    actividad.usuario_email or '',
                    actividad.get_accion_display(),
                    actividad.get_recurso_display(),
                    actividad.recurso_id or '',
                    actividad.ip_address or '',
                    actividad.codigo_estado or '',
                    actividad.duracion_ms or '',
                    actividad.get_nivel_criticidad_display(),
                    json.dumps(actividad.detalle) if actividad.detalle else ''
                ])
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error al exportar CSV: {str(e)}")
            raise
    
    def exportar_excel(self, filtros: Dict = None) -> bytes:
        """
        Exportar logs a Excel
        """
        try:
            actividades = self.repository.obtener_actividades_filtradas(filtros or {})
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Logs"
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Headers
            headers = [
                'ID', 'Fecha', 'Usuario ID', 'Usuario Email', 'Acción',
                'Recurso', 'Recurso ID', 'IP Address', 'Código Estado',
                'Duración (ms)', 'Nivel Criticidad', 'Detalles'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos
            for row, actividad in enumerate(actividades, 2):
                ws.cell(row=row, column=1, value=str(actividad.id))
                ws.cell(row=row, column=2, value=actividad.fecha.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row, column=3, value=actividad.usuario_id or '')
                ws.cell(row=row, column=4, value=actividad.usuario_email or '')
                ws.cell(row=row, column=5, value=actividad.get_accion_display())
                ws.cell(row=row, column=6, value=actividad.get_recurso_display())
                ws.cell(row=row, column=7, value=actividad.recurso_id or '')
                ws.cell(row=row, column=8, value=actividad.ip_address or '')
                ws.cell(row=row, column=9, value=actividad.codigo_estado or '')
                ws.cell(row=row, column=10, value=actividad.duracion_ms or '')
                ws.cell(row=row, column=11, value=actividad.get_nivel_criticidad_display())
                ws.cell(row=row, column=12, value=json.dumps(actividad.detalle) if actividad.detalle else '')
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.read()
            
        except Exception as e:
            logger.error(f"Error al exportar Excel: {str(e)}")
            raise
    
    def exportar_pdf(self, filtros: Dict = None) -> bytes:
        """
        Exportar logs a PDF
        """
        try:
            actividades = self.repository.obtener_actividades_filtradas(filtros or {})
            
            # Crear buffer
            buffer = io.BytesIO()
            
            # Crear documento
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Título
            title = Paragraph("Reporte de Auditoría del Sistema", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Información del reporte
            info_data = [
                ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
                ['Total de registros:', str(len(actividades))],
                ['Microservicio:', 'AUDITORÍA']
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Tabla de datos
            if actividades:
                table_data = [['Fecha', 'Usuario', 'Acción', 'Recurso', 'Estado']]
                
                for actividad in actividades[:100]:  # Limitar a 100 registros
                    table_data.append([
                        actividad.fecha.strftime('%d/%m %H:%M'),
                        actividad.usuario_email[:20] + '...' if actividad.usuario_email and len(actividad.usuario_email) > 20 else (actividad.usuario_email or 'Anónimo'),
                        actividad.get_accion_display()[:15],
                        actividad.get_recurso_display()[:15],
                        str(actividad.codigo_estado) if actividad.codigo_estado else 'N/A'
                    ])
                
                table = Table(table_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.2*inch, 0.8*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (1, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
            else:
                elements.append(Paragraph("No se encontraron registros con los filtros aplicados.", styles['Normal']))
            
            # Generar PDF
            doc.build(elements)
            
            # Obtener datos del buffer
            buffer.seek(0)
            return buffer.read()
            
        except Exception as e:
            logger.error(f"Error al exportar PDF: {str(e)}")
            raise